"""밸류에이션 결과 → Excel 워크북 (수식 기반 + 셀 출처주석)."""
from __future__ import annotations

from openpyxl.utils import get_column_letter

from core.schema import Provenance, SourceType
from excel.workbook import ValuationWorkbook
from engines import sangjeung, dcf, dcf_full, comps

# 법정 파라미터(데이터 아님) 출처
_LAW = Provenance(source="상증세법 시행령 §54 / 시행규칙", source_type=SourceType.REFERENCE,
                  source_url="https://www.law.go.kr", note="법정 계산 파라미터")
_ASSUME = Provenance(source="사용자 가정(입력)", source_type=SourceType.ASSUMPTION,
                     source_url="(assumption)", note="사용자가 입력한 밸류에이션 가정")


def _assume(note: str) -> Provenance:
    return Provenance(source="사용자 가정(입력, 실데이터 역산 기본값 포함)",
                      source_type=SourceType.ASSUMPTION, source_url="(assumption)", note=note)


def sangjeung_workbook(company: str, year: int | None = None,
                       real_estate_heavy: bool = False, report: str = "annual") -> tuple[bytes, str]:
    """상증법 보충적평가 워크북 생성 → (xlsx bytes, 파일명)."""
    m = sangjeung.build_model(company, year, real_estate_heavy, report)
    s = m["ni_series"]
    wb = ValuationWorkbook("상증법 평가")

    wb.title(1, f"상증법 보충적평가 — {m['company']} ({m['as_of']})",
             "⚠️ 세무조정·자산 시가평가 미반영 근사 · 별도재무제표 기준 · 비상장주식 보충적 평가방법")

    wb.section(3, "입력 (DART 전자공시)")
    b_shares = wb.input(4, m["shares"].value, m["shares"].provenance, "발행주식총수 (주)")
    b_ni0 = wb.input(5, s[0]["amount"], m["ni_prov"], f"당기순이익 {s[0]['period']}")
    b_ni1 = wb.input(6, s[1]["amount"], m["ni_prov"], f"당기순이익 {s[1]['period']}")
    b_ni2 = wb.input(7, s[2]["amount"], m["ni_prov"], f"당기순이익 {s[2]['period']}")
    b_eq = wb.input(8, m["equity"].value, m["equity"].provenance, "자본총계 (순자산)")

    wb.section(10, "가정 (상증세법 법정 파라미터)")
    b_cap = wb.input(11, m["cap_rate"], _LAW, "순손익가치 환원율", fmt="0%")
    b_w1 = wb.input(12, m["w_income"], _LAW, "가중치 · 순손익", fmt="0")
    b_w2 = wb.input(13, m["w_asset"], _LAW, "가중치 · 순자산", fmt="0")
    b_floor = wb.input(14, m["floor_pct"], _LAW, "순자산가치 하한율", fmt="0%")

    wb.section(16, "계산 (수식 — 입력을 바꾸면 자동 반영)")
    b_wavg = wb.formula(17, f"=({b_ni0}*3+{b_ni1}*2+{b_ni2})/6", "3개년 가중평균 순손익")
    b_nips = wb.formula(18, f"={b_wavg}/{b_shares}", "1주당 순손익", fmt="#,##0.0")
    b_incv = wb.formula(19, f"=MAX(0,{b_nips}/{b_cap})", "순손익가치 (원/주)")
    b_navps = wb.formula(20, f"={b_eq}/{b_shares}", "순자산가치 (원/주)")
    b_blend = wb.formula(21, f"=({b_incv}*{b_w1}+{b_navps}*{b_w2})/({b_w1}+{b_w2})",
                         f"가중평가액 ({m['w_income']}:{m['w_asset']})")
    b_floorv = wb.formula(22, f"={b_navps}*{b_floor}", "하한 (순자산×80%)")
    wb.formula(23, f"=MAX({b_blend},{b_floorv})", "▶ 1주당 평가액", fmt="#,##0", bold=True)

    wb.note(25, "출처: 각 파란 입력 셀에 마우스를 올리면 DART 공시 출처가 표시됩니다. '출처' 시트 참고.")
    wb.note(26, "한계: 법상 순손익은 '각 사업연도 소득(세무조정)', 순자산은 '시가평가액'이나 "
                "여기서는 공개 재무제표의 당기순이익·자본총계로 근사했습니다.")

    fname = f"상증법_{m['company']}_{m['as_of']}.xlsx".replace("/", "_")
    return wb.to_bytes(), fname


_CURRENCY = {"KR": "KRW", "US": "USD", "JP": "JPY", "TW": "TWD"}


def dcf_workbook(company: str, wacc_pct: float, net_debt: float, revenue_growth,
                 ebit_margin_pct: float, da_pct: float, capex_pct: float, nwc_pct: float,
                 terminal_growth_pct: float, forecast_years: int = 5,
                 tax_rate_pct: float | None = None, year: int | None = None,
                 market: str = "KR") -> tuple[bytes, str]:
    """DCF(UFCF) 워크북 → (xlsx bytes, 파일명). 추정표는 수식, 민감도는 계산 스냅샷."""
    d = dcf.build_model(company, wacc_pct, net_debt, revenue_growth, ebit_margin_pct,
                        da_pct, capex_pct, nwc_pct, terminal_growth_pct,
                        forecast_years, tax_rate_pct, year, market=market)
    N = d["forecast_years"]
    cur = _CURRENCY.get(d["market"], d["market"])
    wb = ValuationWorkbook("DCF")
    wb.title(1, f"DCF 밸류에이션 (UFCF) — {d['company']} ({d['as_of']})",
             f"⚠️ 성장률·마진·Capex 등은 사용자 가정. 기준매출·발행주식수={d['base_revenue'].provenance.source}, 세율=Damodaran.")

    # ── 입력·가정 (B열) ──
    wb.section(3, "입력 · 가정")
    L = lambda r, t: wb.put(r, 1, t)
    L(4, "기준매출 (Year0)");      b_rev0 = wb.icell(4, 2, d["base_revenue"].value, d["base_revenue"].provenance, src_label="기준매출")
    L(5, "발행주식총수");          b_sh = wb.icell(5, 2, d["shares"].value, d["shares"].provenance, src_label="발행주식총수")
    L(6, "법인세율 (%)");         b_tax = wb.icell(6, 2, d["tax_pct"], d["tax_prov"], fmt="0.0", src_label="법인세율")
    L(7, "WACC (%)");           b_wacc = wb.icell(7, 2, d["wacc_pct"], _ASSUME, fmt="0.00", src_label="WACC")
    L(8, "Terminal g (%)");     b_g = wb.icell(8, 2, d["terminal_growth_pct"], _ASSUME, fmt="0.00", src_label="Terminal g")
    L(9, "순부채 (차입금−현금)");    b_nd = wb.icell(9, 2, d["net_debt"], _ASSUME, src_label="순부채")
    L(10, "영업이익률 (%)");       b_m = wb.icell(10, 2, d["ebit_margin_pct"], _ASSUME, fmt="0.0", src_label="영업이익률")
    L(11, "D&A (% 매출)");       b_da = wb.icell(11, 2, d["da_pct"], _ASSUME, fmt="0.0", src_label="D&A%")
    L(12, "Capex (% 매출)");     b_cx = wb.icell(12, 2, d["capex_pct"], _ASSUME, fmt="0.0", src_label="Capex%")
    L(13, "ΔNWC (% 매출증가)");   b_nwc = wb.icell(13, 2, d["nwc_pct"], _ASSUME, fmt="0.0", src_label="ΔNWC%")

    # ── 추정표 (연도=열, 수식) ──
    wb.section(15, "추정 (수식 — 가정 바꾸면 자동 flex)")
    hdr = 16
    wb.put(hdr, 1, "연차", bold=True)
    wb.put(hdr, 2, "기준(0)", bold=True)
    cols = [get_column_letter(3 + i) for i in range(N)]     # C, D, E, ...
    for i, cl in enumerate(cols, 1):
        wb.fcell(hdr, 2 + i, i, fmt="0")                    # year index (정수) in row 16
    # 성장률(입력)
    wb.put(17, 1, "매출성장률 (%)")
    for i, cl in enumerate(cols):
        wb.icell(17, 3 + i, d["growth_pct"][i], _ASSUME, fmt="0.0", src_label=f"성장률 Y{i+1}")
    # 매출
    wb.put(18, 1, "매출")
    wb.fcell(18, 2, f"={b_rev0}")
    for i, cl in enumerate(cols):
        prev = get_column_letter(2 + i)                     # B, C, D...
        wb.fcell(18, 3 + i, f"={prev}18*(1+{cl}17/100)")
    # EBIT / NOPAT / D&A / Capex / ΔNWC / UFCF / DF / PV
    def rowf(r, label, fn, fmt="#,##0"):
        wb.put(r, 1, label)
        for i, cl in enumerate(cols):
            wb.fcell(r, 3 + i, fn(cl, get_column_letter(2 + i)), fmt=fmt)
    rowf(19, "EBIT", lambda cl, pv: f"={cl}18*$B$10/100")
    rowf(20, "세후영업이익(NOPAT)", lambda cl, pv: f"={cl}19*(1-$B$6/100)")
    rowf(21, "(+) D&A", lambda cl, pv: f"={cl}18*$B$11/100")
    rowf(22, "(−) Capex", lambda cl, pv: f"={cl}18*$B$12/100")
    rowf(23, "(−) ΔNWC", lambda cl, pv: f"=({cl}18-{pv}18)*$B$13/100")
    rowf(24, "= UFCF", lambda cl, pv: f"={cl}20+{cl}21-{cl}22-{cl}23")
    rowf(25, "할인계수", lambda cl, pv: f"=1/(1+$B$7/100)^{cl}16", fmt="0.000")
    rowf(26, "PV(UFCF)", lambda cl, pv: f"={cl}24*{cl}25")

    last = cols[-1]
    wb.section(28, "가치평가 (수식)")
    wb.put(29, 1, "PV(UFCF) 합");        b_pvsum = wb.fcell(29, 2, f"=SUM(C26:{last}26)")
    wb.put(30, 1, "Terminal Value");     b_tv = wb.fcell(30, 2, f"={last}24*(1+$B$8/100)/($B$7/100-$B$8/100)")
    wb.put(31, 1, "PV(Terminal Value)"); b_pvtv = wb.fcell(31, 2, f"={b_tv}/(1+$B$7/100)^{last}16")
    wb.put(32, 1, "기업가치 (EV)");       b_ev = wb.fcell(32, 2, f"={b_pvsum}+{b_pvtv}", bold=True)
    wb.put(33, 1, "(−) 순부채");          wb.fcell(33, 2, f"={b_nd}")
    wb.put(34, 1, "지분가치");            b_eq = wb.fcell(34, 2, f"={b_ev}-{b_nd}")
    wb.put(35, 1, f"▶ 주당가치 ({cur})");  wb.fcell(35, 2, f"={b_eq}/{b_sh}", bold=True)

    # ── [검증 체크] — dcf_full_workbook 과 동일한 패턴(WACC/g·TV 비중) ──
    wb.section(37, "[검증 체크 — Validation Checks]")
    wb.put(38, 1, "WACC > g ?")
    wb.fcell(38, 2, f'=IF({b_wacc}<={b_g},"⚠ FAIL: WACC≤g — TV 무효","OK")')
    wb.put(39, 1, "Terminal UFCF > 0 ?")
    wb.fcell(39, 2, f'=IF({last}24<=0,"⚠ Terminal UFCF≤0 — Gordon Growth 신뢰불가","OK")')
    wb.put(40, 1, "TV 비중 (PV(TV)/EV)")
    wb.fcell(40, 2, f'=IF({b_ev}>0,{b_pvtv}/{b_ev},"n/a")', fmt="0.0%")
    wb.put(41, 1, "TV 비중 평가")
    wb.fcell(41, 2, '=IF(ISNUMBER(B40),IF(B40>0.9,"⚠ High-risk: TV>90%",'
                    'IF(B40>0.75,"⚠ Warning: TV>75%","OK")),"n/a")')

    # ── 민감도 (계산 스냅샷) ──
    wb.section(43, "민감도: 주당가치 (WACC × Terminal g) · 계산 스냅샷")
    wb.put(44, 1, "WACC\\g", bold=True)
    wacc_axis = [d["wacc_pct"] + dv for dv in (-1, -0.5, 0, 0.5, 1)]
    g_axis = [d["terminal_growth_pct"] + dv for dv in (-1, -0.5, 0, 0.5, 1)]
    for j, g in enumerate(g_axis, 2):
        wb.put(44, j, f"{g:.1f}%", bold=True)
    for r, w in enumerate(wacc_axis, 45):
        wb.put(r, 1, f"{w:.2f}%", bold=True)
        for j, g in enumerate(g_axis, 2):
            try:
                s = dcf.build_model(company, w, net_debt, revenue_growth, ebit_margin_pct,
                                    da_pct, capex_pct, nwc_pct, g, forecast_years,
                                    tax_rate_pct, year, market=market)
                wb.ws.cell(r, j, round(s["per_share"])).number_format = "#,##0"
            except Exception:  # noqa: BLE001 (WACC<=g 등)
                wb.ws.cell(r, j, "n/a")

    wb.note(51, "민감도표는 현재 가정 기준 계산 스냅샷입니다. 위 추정표는 입력을 바꾸면 실시간 flex 됩니다.")
    wb.note(52, "한계: 순부채·성장·마진·Capex 등은 사용자 가정이며, 결과는 가정에 매우 민감합니다.")

    fname = f"DCF_{d['company']}_{d['as_of']}.xlsx".replace("/", "_")
    return wb.to_bytes(), fname


def comps_workbook(target: str, peers: list, year: int | None = None) -> tuple[bytes, str]:
    """Trading comps 워크북 → (xlsx bytes, 파일명). 배수·적용은 수식."""
    m = comps.build_model(target, peers, year)
    wb = ValuationWorkbook("Comps")
    wb.title(1, f"Trading Comps (자기자본배수) — {m['target']} ({m['as_of']})",
             "⚠️ PER·PBR 배수. 시가총액=네이버(KRX), 순이익·자본총계=DART. EV배수는 추후.")

    wb.section(3, "Peer 배수 (PER=시총/순이익, PBR=시총/자본)")
    hdr = 4
    for c, t in enumerate(["회사", "시가총액", "당기순이익", "자본총계", "PER(x)", "PBR(x)"], 1):
        wb.put(hdr, c, t, bold=True)
    mc_prov = Provenance(source="네이버 금융(KRX 시세)", source_type=SourceType.AUTHORITATIVE,
                         source_url="https://m.stock.naver.com", original_field="시가총액")
    dart_prov = Provenance(source="DART (금융감독원)", source_type=SourceType.AUTHORITATIVE,
                           source_url="https://dart.fss.or.kr", original_field="당기순이익·자본총계")
    r = hdr + 1
    first = r
    for p in m["peers"]:
        wb.put(r, 1, f"{p['name']} ({p['stock_code']})")
        wb.icell(r, 2, p["market_cap"], mc_prov, src_label=f"{p['name']} 시총")
        wb.icell(r, 3, p["net_income"], dart_prov, src_label=f"{p['name']} 순이익")
        wb.icell(r, 4, p["equity"], dart_prov, src_label=f"{p['name']} 자본")
        if p["per"]:
            wb.fcell(r, 5, f"=B{r}/C{r}", fmt="0.0")
        else:
            wb.put(r, 5, "n/a")
        if p["pbr"]:
            wb.fcell(r, 6, f"=B{r}/D{r}", fmt="0.00")
        else:
            wb.put(r, 6, "n/a")
        r += 1
    last = r - 1
    wb.put(r, 1, "중앙값 (median)", bold=True)
    b_per = wb.fcell(r, 5, f"=MEDIAN(E{first}:E{last})", fmt="0.0", bold=True)
    b_pbr = wb.fcell(r, 6, f"=MEDIAN(F{first}:F{last})", fmt="0.00", bold=True)
    med_row = r

    r += 2
    wb.section(r, "타깃 적용 → 주당가치")
    r += 1
    wb.put(r, 1, f"타깃: {m['target']}")
    r += 1
    wb.put(r, 1, "당기순이익")
    b_ni = wb.icell(r, 2, m["ni_t"].value, m["ni_t"].provenance, src_label="타깃 순이익")
    r += 1
    wb.put(r, 1, "자본총계")
    b_eq = wb.icell(r, 2, m["eq_t"].value, m["eq_t"].provenance, src_label="타깃 자본")
    r += 1
    wb.put(r, 1, "발행주식총수")
    b_sh = wb.icell(r, 2, m["shares_t"].value, m["shares_t"].provenance, src_label="타깃 주식수")
    r += 1
    wb.put(r, 1, "▶ 주당가치 (PER 기준)")
    wb.fcell(r, 2, f"={b_ni}*{b_per}/{b_sh}", bold=True)
    r += 1
    wb.put(r, 1, "▶ 주당가치 (PBR 기준)")
    wb.fcell(r, 2, f"={b_eq}*{b_pbr}/{b_sh}", bold=True)

    if m["errors"]:
        r += 2
        wb.note(r, "제외된 peer: " + " / ".join(m["errors"]))
    r += 1
    wb.note(r, "PER 기준은 이익, PBR 기준은 순자산 관점. 두 값의 범위를 함께 보세요. 배수 셀은 수식이라 입력 변경 시 flex.")

    fname = f"Comps_{m['target']}_{m['as_of']}.xlsx".replace("/", "_")
    return wb.to_bytes(), fname


def _hist_prov(m: dict, label: str) -> Provenance:
    src = m.get("sources", {}).get("revenue") or {}
    rcept = src.get("rcept")
    return Provenance(
        source="DART (금융감독원)", source_type=SourceType.AUTHORITATIVE,
        source_url=(f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept}" if rcept
                   else "https://dart.fss.or.kr"),
        original_field=label, note="5개년 각 사업연도 정기보고서(fnlttSinglAcntAll) 기준",
    )


def dcf_full_workbook(company: str, wacc_pct: float, net_debt: float, revenue_growth_pct,
                      ebit_margin_pct: float, da_pct: float, capex_pct: float,
                      nwc_pct: float, terminal_growth_pct: float, forecast_years: int = 5,
                      tax_rate_pct: float | None = None, year: int | None = None,
                      fcff_mode: str = "raw") -> tuple[bytes, str]:
    """5시트 통합 DCF 워크북: P&L and FCF / Debt Schedule / Depreciation Schedule /
    DCF Valuation / Assumption Summary. compute_dcf 챗 tool 의 입력을 그대로 받는다(파라미터명도
    동일하게 revenue_growth_pct — registry.py 의 _dcf 와 일치시켜야 앱에서 그대로 재사용 가능).
    net_debt/da_pct 는 이 모델에선 실데이터(차입금-현금)·Depreciation Schedule 이 대신하므로
    미사용(호출부 호환을 위해 인자만 유지). fcff_mode="raw"(기본) 는 DCF Valuation 시트의
    헤드라인 계산에 zero-floor 되지 않은 FCFF 를 쓴다 — zero-floor 는 시트 내 대안 시나리오로
    항상 함께 표시된다(값을 숨기지 않는다는 원칙)."""
    assumptions = {
        "wacc_pct": wacc_pct, "revenue_growth": revenue_growth_pct,
        "ebit_margin_pct": ebit_margin_pct, "capex_pct": capex_pct, "nwc_pct": nwc_pct,
        "terminal_growth_pct": terminal_growth_pct, "forecast_years": forecast_years,
        "tax_rate_pct": tax_rate_pct, "fcff_mode": fcff_mode,
    }
    m = dcf_full.build_full_model(company, assumptions, year)
    n_hist, n_fore = len(m["hist_years"]), m["n_fore"]
    col = lambda i: get_column_letter(2 + i)  # B=index0
    hist_cols = [col(i) for i in range(n_hist)]
    fore_cols = [col(n_hist + i) for i in range(n_fore)]
    all_years = m["hist_years"] + m["fore_years"]
    all_cols = hist_cols + fore_cols

    wb = ValuationWorkbook("P&L and FCF")
    wb.title(1, f"{m['company']} DCF Valuation", "⚠️ 예측·가정 구간(파란 글씨)은 조정 가능 — 셀 주석에 출처/근거 표시")
    wb.put(2, 1, "P&L and FCF (단위: 원)", bold=True)
    wb.put(3, 1, "구분", bold=True)
    for c, yr in zip(hist_cols, m["hist_years"]):
        wb.put(3, ord(c) - 64, str(yr), bold=True)
    for c, yr in zip(fore_cols, m["fore_years"]):
        wb.put(3, ord(c) - 64, f"{yr}E", bold=True)

    pl = m["pl"]

    def _row_hist_input(row: int, label: str, values: list, prov_label: str, unit_note: str = ""):
        wb.put(row, 1, label)
        cells = []
        for c, v in zip(hist_cols, values):
            if v is None:
                continue
            cells.append(wb.icell(row, ord(c) - 64, round(v), _hist_prov(m, prov_label),
                                  src_label=f"{label}{unit_note}"))
        return cells

    _row_hist_input(4, "매출액", pl["rev_hist"], "매출액")
    for c, v in zip(fore_cols, pl["rev_fore"]):
        wb.fcell(4, ord(c) - 64, f"={_prev_col(c)}4*(1+{c}42/100)")
    _row_hist_input(5, "매출원가", pl["cogs_hist"], "매출원가")
    for c in fore_cols:
        wb.fcell(5, ord(c) - 64, f"={c}4*$B$40/100")
    for c in all_cols:
        wb.fcell(6, ord(c) - 64, f"={c}4-{c}5")
    wb.put(6, 1, "매출총이익")
    for c in all_cols:
        wb.fcell(7, ord(c) - 64, f"={c}6/{c}4", fmt="0.0%")
    wb.put(7, 1, "매출총이익률(%)")
    _row_hist_input(8, "판매비와관리비", pl["sga_hist"], "판매비와관리비")
    for c in fore_cols:
        wb.fcell(8, ord(c) - 64, f"={c}4*$B$41/100")
    for c in all_cols:
        wb.fcell(9, ord(c) - 64, f"={c}6-{c}8")
    wb.put(9, 1, "영업이익")
    for c in all_cols:
        wb.fcell(10, ord(c) - 64, f"={c}9/{c}4", fmt="0.0%")
    wb.put(10, 1, "영업이익률(%)")
    _row_hist_input(11, "이자비용(금융비용)", pl["ie_hist"], "이자비용")
    for c, yr in zip(fore_cols, m["fore_years"]):
        wb.fcell(11, ord(c) - 64, f"={wb.qref('Debt Schedule', f'{_debt_col(m, yr)}23')}")
    for c in all_cols:
        wb.fcell(12, ord(c) - 64, f"={c}9-{c}11")
    wb.put(12, 1, "법인세비용차감전이익")
    _row_hist_input(13, "법인세비용", pl["tax_hist"], "법인세비용")
    for c in fore_cols:
        wb.fcell(13, ord(c) - 64, f"=MAX(0,{c}12*$B$44/100)")
    _row_hist_input(14, "당기순이익", pl["ni_hist"], "당기순이익")
    for c in fore_cols:
        wb.fcell(14, ord(c) - 64, f"={c}12-{c}13")
    for c in all_cols:
        wb.fcell(15, ord(c) - 64, f"={c}14/{c}4", fmt="0.0%")
    wb.put(15, 1, "순이익률(%)")

    _row_hist_input(17, "감가상각비(D&A)", pl["da_hist"], "감가상각비")
    for c, yr in zip(fore_cols, m["fore_years"]):
        wb.fcell(17, ord(c) - 64, f"={wb.qref('Depreciation Schedule', f'{_dep_col(m, yr)}20')}")
    for c in all_cols:
        wb.fcell(18, ord(c) - 64, f"={c}9+{c}17")
    wb.put(18, 1, "EBITDA")
    _row_hist_input(19, "CAPEX", pl["capex_hist"], "CAPEX(유형자산의 취득)")
    for c in fore_cols:
        wb.fcell(19, ord(c) - 64, f"={c}4*$B$42/100")
    _row_hist_input(20, "운전자본 변동(ΔWC)", pl["nwc_hist"], "운전자본 변동")
    for i, c in enumerate(fore_cols):
        prev = _prev_col(c)
        wb.fcell(20, ord(c) - 64, f"=({c}4-{prev}4)*$B$43/100")
    wb.put(20, 1, "운전자본 변동(ΔWC)")

    for c in all_cols:
        wb.fcell(22, ord(c) - 64, f"={c}9*(1-$B$44/100)")
    wb.put(22, 1, "NOPAT")
    for c in all_cols:
        wb.fcell(23, ord(c) - 64, f"={c}22+{c}17-{c}19-{c}20")
    wb.put(23, 1, "FCFF (Raw)")
    # zero-floor 는 예측 구간에 대한 대안 시나리오일 뿐 — 실적(historical) 열은 사실이므로
    # floor 하지 않는다(마이너스 실제 현금흐름을 0으로 왜곡하지 않음).
    for c in fore_cols:
        wb.fcell(24, ord(c) - 64, f"=MAX(0,{c}23)", bold=True)
    for c in hist_cols:
        wb.put(24, ord(c) - 64, "–")
    wb.put(24, 1, "FCFF (Zero-floor 시나리오, 예측만)", bold=True)

    _row_hist_input(26, "영업활동현금흐름", pl["ocf_hist"], "영업활동현금흐름")
    for c in all_cols:
        if c in hist_cols:
            wb.fcell(27, ord(c) - 64, f"={c}26-{c}19")
    wb.put(27, 1, "검증 FCFF (OCF−CAPEX)")

    wb.note(29, "히스토리(실선 열)는 DART 실측값(주석에 출처). 예측(E 표시)은 아래 [가정] 셀 기반 수식 — "
                "가정을 바꾸면 이 시트가 자동으로 다시 계산됩니다.")
    if any(v is None for v in pl["da_hist"]):
        wb.note(30, "⚠️ 이 회사는 감가상각비를 현금흐름표에서 구조화 데이터로 찾지 못해 "
                    "EBITDA·FCFF 히스토리가 비어있거나 근사치입니다(D&A=0으로 계산될 수 있음).")

    wb.section(39, "[가정] (파란 셀 — 조정하면 위 표가 자동 반영)")
    a = m["assumptions"]
    wb.icell(40, 2, round(a["cogs_pct"], 2), _assume("최근 3개년 매출원가/매출 평균(실데이터 역산 기본값)"),
            fmt="0.00", src_label="매출원가율(%)")
    wb.put(40, 1, "매출원가율(%)")
    wb.icell(41, 2, round(a["sga_pct"], 2), _assume("최근 3개년 판관비/매출 평균(실데이터 역산 기본값)"),
            fmt="0.00", src_label="판관비율(%)")
    wb.put(41, 1, "판관비율(%)")
    wb.icell(42, 2, round(capex_pct, 2), _assume("사용자 입력(compute_dcf)"), fmt="0.00",
            src_label="Capex율(%)")
    wb.put(42, 1, "CAPEX율(%, 매출대비)")
    wb.icell(43, 2, round(nwc_pct, 2), _assume("사용자 입력(compute_dcf)"), fmt="0.00",
            src_label="ΔWC율(%)")
    wb.put(43, 1, "ΔWC율(%, 매출증가대비)")
    wb.icell(44, 2, round(a["tax_rate_pct"], 2), _assume("Damodaran 한국 법인세율(미지정 시)"),
            fmt="0.00", src_label="법인세율(%)")
    wb.put(44, 1, "법인세율(%)")
    for i, (c, g) in enumerate(zip(fore_cols, a["growth"])):
        wb.icell(45, ord(c) - 64, round(g, 2), _assume("최근 CAGR 기반 완만한 감소(사용자 입력 없을 때 기본값)"),
                fmt="0.00", src_label=f"매출성장률 {m['fore_years'][i]}(%)")
    wb.put(45, 1, "매출성장률(%, 연차별)")

    # ── 시트2: Debt Schedule ──
    wb.new_sheet("Debt Schedule")
    d = m["debt"]
    wb.title(1, f"{m['company']} — Debt Schedule", "단위: 원. 개별 대출건 금리·상환계획은 구조화 데이터가 없어 가정(파란 셀)")
    wb.put(3, 1, "구분"); wb.put(3, 2, "이자율")
    d_base_col = "C"
    d_fore_cols = [get_column_letter(4 + i) for i in range(n_fore)]
    wb.put(3, 3, str(all_years[n_hist - 1]), bold=True)
    for c, yr in zip(d_fore_cols, m["fore_years"]):
        wb.put(3, ord(c) - 64, f"{yr}E", bold=True)

    wb.section(4, "[단기차입금]")
    wb.put(5, 1, "기초잔액")
    wb.icell(5, ord(d_fore_cols[0]) - 64, round(d["short0"]),
            m["sources"]["short_debt"].provenance, src_label="단기차입금 기초잔액")
    for i in range(1, n_fore):
        prev = d_fore_cols[i - 1]
        wb.fcell(5, ord(d_fore_cols[i]) - 64, f"={prev}9")
    wb.put(6, 1, "신규차입")
    for c in d_fore_cols:
        wb.icell(6, ord(c) - 64, 0, _assume("기본값: 신규차입 없음(가정)"), src_label="단기 신규차입")
    wb.put(7, 1, "상환")
    for c in d_fore_cols:
        wb.icell(7, ord(c) - 64, -round(d["short_repay"][0]) if d["short0"] else 0,
                _assume("기본값: 단기차입금 기초잔액의 약 8%/년(가정 — 실제 만기·상환계획으로 조정)"),
                src_label="단기 상환")
    wb.put(8, 1, "환율조정")
    for c in d_fore_cols:
        wb.icell(8, ord(c) - 64, 0, _assume("기본값 0(외화차입 없다고 가정)"), src_label="환율조정")
    for c in d_fore_cols:
        wb.fcell(9, ord(c) - 64, f"={c}5+{c}6+{c}7+{c}8")
    wb.put(9, 1, "기말잔액", bold=True)

    wb.section(11, "[장기차입금]")
    wb.put(12, 1, "기초잔액")
    wb.icell(12, ord(d_fore_cols[0]) - 64, round(d["long0"]),
            m["sources"]["long_debt"].provenance, src_label="장기차입금 기초잔액")
    for i in range(1, n_fore):
        prev = d_fore_cols[i - 1]
        wb.fcell(12, ord(d_fore_cols[i]) - 64, f"={prev}15")
    wb.put(13, 1, "신규차입")
    for c in d_fore_cols:
        wb.icell(13, ord(c) - 64, 0, _assume("기본값: 신규차입 없음(가정)"), src_label="장기 신규차입")
    wb.put(14, 1, "상환")
    for c in d_fore_cols:
        wb.icell(14, ord(c) - 64, -round(d["long_repay"][0]) if d["long0"] else 0,
                _assume("기본값: 장기차입금 기초잔액의 약 6%/년(가정 — 실제 만기·상환계획으로 조정)"),
                src_label="장기 상환")
    for c in d_fore_cols:
        wb.fcell(15, ord(c) - 64, f"={c}12+{c}13+{c}14")
    wb.put(15, 1, "기말잔액", bold=True)

    wb.section(17, "[요약]")
    wb.put(18, 1, "총차입금", bold=True)
    wb.fcell(18, 3, f"={d_fore_cols[0]}5+{d_fore_cols[0]}12", bold=True)
    for c in d_fore_cols:
        wb.fcell(18, ord(c) - 64, f"={c}9+{c}15", bold=True)

    wb.put(21, 1, "단기이자비용")
    wb.icell(21, 2, round(d["short_rate_pct"], 2), _assume(
        "실측 역산: 최근 이자비용÷평균 단기차입금(가능하면), 아니면 시중 단기대출금리 참고 기본값"),
        fmt="0.00", src_label="단기 이자율(%)")
    for c in d_fore_cols:
        wb.fcell(21, ord(c) - 64, f"=AVERAGE({c}5,{c}9)*$B$21/100")
    wb.put(22, 1, "장기이자비용")
    wb.icell(22, 2, round(d["long_rate_pct"], 2), _assume("시중 장기대출금리 참고 기본값(조정 필요)"),
            fmt="0.00", src_label="장기 이자율(%)")
    for c in d_fore_cols:
        wb.fcell(22, ord(c) - 64, f"=AVERAGE({c}12,{c}15)*$B$22/100")
    wb.put(23, 1, "총이자비용", bold=True)
    for c in d_fore_cols:
        wb.fcell(23, ord(c) - 64, f"={c}21+{c}22", bold=True)

    wb.note(25, "※ 이자비용은 'P&L and FCF' 시트의 금융비용(예측 구간)과 연동됩니다.")
    wb.note(26, "※ 개별 대출건 금리·실제 상환계획은 DART 구조화 데이터에 없어 위 이자율·상환액은 "
                "가정입니다(파란 셀) — 실제 조건을 알면 직접 조정하세요.")

    # ── 시트3: Depreciation Schedule ──
    wb.new_sheet("Depreciation Schedule")
    dep = m["depreciation"]
    e_fore_cols = [get_column_letter(4 + i) for i in range(n_fore)]
    e_base_col = "C"
    wb.title(1, f"{m['company']} — Depreciation Schedule",
             "단위: 원. 자산별 상각내역은 구조화 데이터가 없어 2버킷(건물/기계장치) 가정으로 근사")
    wb.put(3, 1, "구분"); wb.put(3, 2, "내용연수")
    wb.put(3, 3, f"{all_years[n_hist-1]}년말", bold=True)
    for c, yr in zip(e_fore_cols, m["fore_years"]):
        wb.put(3, ord(c) - 64, f"{yr}E", bold=True)

    wb.section(4, "[기존 자산 감가상각]")
    wb.put(5, 1, f"{all_years[n_hist-1]}년말 유형자산 장부가액")
    wb.icell(5, 3, round(dep["ppe0"]), m["sources"]["ppe"].provenance, src_label="유형자산 장부가액")
    wb.put(6, 1, "기존 자산 감가상각(건물)")
    wb.icell(6, 2, dep["building_life"], _assume("가정(조정 필요) — 실제 자산 구성 모름"), src_label="건물 내용연수")
    for c in e_fore_cols:
        wb.fcell(6, ord(c) - 64, f"=$C$5*$D$8/100/$B$6")
    wb.put(7, 1, "기존 자산 감가상각(기계장치)")
    wb.icell(7, 2, dep["machinery_life"], _assume("가정(조정 필요) — 실제 자산 구성 모름"), src_label="기계장치 내용연수")
    for c in e_fore_cols:
        wb.fcell(7, ord(c) - 64, f"=$C$5*(1-$D$8/100)/$B$7")
    wb.icell(8, 4, round(dep["building_pct"], 1), _assume("가정(조정 필요) — 건물 비중, 기본값 60%"),
            fmt="0.0", src_label="건물 비중(%)")
    wb.put(8, 3, "건물 비중(%) →")
    for c in e_fore_cols:
        wb.fcell(9, ord(c) - 64, f"={c}6+{c}7")
    wb.put(9, 1, "기존 자산 감가상각 소계", bold=True)

    wb.section(11, "[신규 CAPEX 감가상각]")
    wb.put(12, 1, "평균 상각 내용연수")
    wb.icell(12, 2, dep["new_capex_life"],
            _assume("실측 역산(유형자산/최근 D&A, D&A 없으면 10년 기본값)"), src_label="신규 CAPEX 내용연수")
    wb.put(13, 1, "당해연도 CAPEX (참고, 상각 미포함)")
    for c, yr in zip(e_fore_cols, m["fore_years"]):
        wb.fcell(13, ord(c) - 64, f"={wb.qref('P&L and FCF', f'{_pl_col(m, yr)}19')}")
    vintage_start = 14
    for i, yr in enumerate(m["fore_years"]):
        row = vintage_start + i
        wb.put(row, 1, f"{yr}년 CAPEX 상각")
        for j, c in enumerate(e_fore_cols):
            if j <= i:
                continue  # 발생연도 자신은 상각 0(참고 행에 원금이 별도 표시됨)
            life_ok = (j - i) <= dep["new_capex_life"]
            if life_ok:
                wb.fcell(row, ord(c) - 64, f"={e_fore_cols[i]}13/$B$12")
    subtotal_row = vintage_start + n_fore
    for c in e_fore_cols:
        wb.fcell(subtotal_row, ord(c) - 64,
                f"=SUM({c}{vintage_start}:{c}{subtotal_row - 1})")
    wb.put(subtotal_row, 1, "신규 CAPEX 감가상각 소계", bold=True)

    total_row = subtotal_row + 3
    wb.section(total_row - 1, "[총 감가상각비]")
    for c in e_fore_cols:
        wb.fcell(total_row, ord(c) - 64, f"={c}9+{c}{subtotal_row}", bold=True)
    wb.put(total_row, 1, "총 감가상각비", bold=True)
    wb.note(total_row + 2, "※ 총 감가상각비는 'P&L and FCF' 시트의 D&A(예측 구간)와 연동됩니다.")
    wb.note(total_row + 3, "※ 자산별(건물/기계장치) 실제 비중·상각내역은 구조화 데이터가 없어 가정입니다 — "
                          "실제 고정자산 명세가 있으면 직접 조정하세요.")

    # ── 시트4: DCF Valuation ──
    wb.new_sheet("DCF Valuation")
    dcf_m = m["dcf"]["primary"]
    dcf_mode = m["dcf"]["mode"]
    f_fore_cols = [get_column_letter(3 + i) for i in range(n_fore)]
    f_term_col = get_column_letter(3 + n_fore)
    wb.title(1, f"{m['company']} — DCF Valuation",
             f"단위: 원 · 헤드라인 기준 FCFF = {dcf_mode} (기본은 raw — zero-floor는 대안 시나리오로 병기)")
    wb.section(3, "[DCF 가정]")
    wb.icell(4, 2, wacc_pct, _assume("사용자 입력(compute_dcf)"), fmt="0.00", src_label="WACC(%)")
    wb.put(4, 1, "WACC(%)")
    wb.icell(5, 2, terminal_growth_pct, _assume("사용자 입력(compute_dcf)"), fmt="0.00",
            src_label="영구성장률(%)")
    wb.put(5, 1, "영구성장률(g, %)")
    wb.icell(6, 2, dcf_m["shares"], m["sources"]["shares"].provenance, fmt="#,##0",
            src_label="발행주식수")
    wb.put(6, 1, "발행주식수")
    # 순부채 = (단기+장기 차입금 기초) - 현금. 차입금은 Debt Schedule 시트 참조, 현금은 이 시트에 직접 입력.
    wb.icell(7, 2, round(m["sources"]["cash"].value), m["sources"]["cash"].provenance,
            src_label="현금및현금성자산")
    wb.put(7, 1, "현금및현금성자산")
    wb.fcell(8, 2, f"=({wb.qref('Debt Schedule', f'{d_fore_cols[0]}5')}+"
                  f"{wb.qref('Debt Schedule', f'{d_fore_cols[0]}12')})-B7", bold=True)
    wb.put(8, 1, "순부채", bold=True)

    headline_pl_row = 23 if dcf_mode == "raw" else 24  # P&L 시트: 23=Raw FCFF, 24=Zero-floor
    alt_pl_row = 24 if headline_pl_row == 23 else 23
    alt_mode = "zero_floor" if dcf_mode == "raw" else "raw"

    wb.section(11, "[FCFF 및 현재가치]")
    wb.put(12, 1, "구분")
    for c, yr in zip(f_fore_cols, m["fore_years"]):
        wb.put(12, ord(c) - 64, f"{yr}E", bold=True)
    wb.put(12, ord(f_term_col) - 64, "Terminal", bold=True)
    wb.put(13, 1, f"FCFF ({dcf_mode}, 기본)")
    for c, yr in zip(f_fore_cols, m["fore_years"]):
        wb.fcell(13, ord(c) - 64, f"={wb.qref('P&L and FCF', f'{_pl_col(m, yr)}{headline_pl_row}')}")
    wb.put(14, 1, "할인기간(년)")
    for i, c in enumerate(f_fore_cols, 1):
        wb.put(14, ord(c) - 64, i)
    wb.put(15, 1, "할인계수")
    for c in f_fore_cols:
        wb.fcell(15, ord(c) - 64, f"=1/(1+$B$4/100)^{c}14", fmt="0.0000")
    wb.put(16, 1, "현재가치(PV)")
    for c in f_fore_cols:
        wb.fcell(16, ord(c) - 64, f"={c}13*{c}15")

    last_col = f_fore_cols[-1]
    wb.section(18, "[Terminal Value 계산]")
    wb.put(19, 1, f"{m['fore_years'][-1]} FCFF×(1+g)")
    wb.fcell(19, ord(f_term_col) - 64, f"={last_col}13*(1+$B$5/100)")
    wb.put(20, 1, "Terminal Value")
    wb.fcell(20, ord(f_term_col) - 64, f"={f_term_col}19/($B$4/100-$B$5/100)")
    wb.put(21, 1, "TV 현재가치")
    wb.fcell(21, ord(f_term_col) - 64, f"={f_term_col}20*{last_col}15")

    wb.section(23, "[기업가치 및 주주가치]")
    wb.put(24, 1, "PV(FCFF 합계)")
    wb.fcell(24, 2, f"=SUM({f_fore_cols[0]}16:{last_col}16)")
    wb.put(25, 1, "PV(Terminal Value)")
    wb.fcell(25, 2, f"={f_term_col}21")
    wb.put(26, 1, "기업가치(EV)", bold=True)
    wb.fcell(26, 2, "=B24+B25", bold=True)
    wb.put(27, 1, "(−) 순부채")
    wb.fcell(27, 2, "=B8")
    wb.put(28, 1, "지분가치")
    wb.fcell(28, 2, "=B26-B27")
    wb.put(29, 1, "▶ 주당가치 (원)", bold=True)
    wb.fcell(29, 2, '=IF(B28<=0,"NM",B28/B6)', bold=True)

    # ── [검증 체크] — 스펙 N/S: TV·WACC/g·지분가치 이상 여부를 formula 로 상시 감시 ──
    wb.section(31, "[검증 체크 — Validation Checks]")
    wb.put(32, 1, "WACC > g ?")
    wb.fcell(32, 2, '=IF(B4<=B5,"⚠ FAIL: WACC≤g — TV 무효","OK")')
    wb.put(33, 1, "Terminal FCFF > 0 ?")
    wb.fcell(33, 2, f'=IF({f_term_col}19<=0,"⚠ Terminal FCFF≤0 — Gordon Growth 신뢰불가","OK")')
    wb.put(34, 1, "TV 비중 (PV(TV)/EV)")
    wb.fcell(34, 2, '=IF(B26>0,B25/B26,"n/a")', fmt="0.0%")
    wb.put(35, 1, "TV 비중 평가")
    wb.fcell(35, 2, '=IF(ISNUMBER(B34),IF(B34>0.9,"⚠ High-risk: TV>90%",'
                    'IF(B34>0.75,"⚠ Warning: TV>75%","OK")),"n/a")')
    wb.put(36, 1, "지분가치 음수 체크")
    wb.fcell(36, 2, '=IF(B28<0,"⚠ 지분가치 음수 — 주당가치 NM(B28 원본값 참고)","OK")')

    # ── [대안 시나리오] — 헤드라인이 아닌 FCFF 방식(raw↔zero_floor)을 항상 병기해 값을 숨기지 않음 ──
    wb.section(38, f"[대안 시나리오: FCFF({alt_mode}) — 참고용, 기본은 {dcf_mode}]")
    wb.put(39, 1, "구분")
    for c, yr in zip(f_fore_cols, m["fore_years"]):
        wb.put(39, ord(c) - 64, f"{yr}E", bold=True)
    wb.put(39, ord(f_term_col) - 64, "Terminal", bold=True)
    wb.put(40, 1, f"FCFF ({alt_mode})")
    for c, yr in zip(f_fore_cols, m["fore_years"]):
        wb.fcell(40, ord(c) - 64, f"={wb.qref('P&L and FCF', f'{_pl_col(m, yr)}{alt_pl_row}')}")
    wb.put(41, 1, "현재가치(PV)")
    for c in f_fore_cols:
        wb.fcell(41, ord(c) - 64, f"={c}40*{c}15")
    wb.put(42, 1, f"{m['fore_years'][-1]} FCFF×(1+g)")
    wb.fcell(42, ord(f_term_col) - 64, f"={last_col}40*(1+$B$5/100)")
    wb.put(43, 1, "Terminal Value")
    wb.fcell(43, ord(f_term_col) - 64, f"={f_term_col}42/($B$4/100-$B$5/100)")
    wb.put(44, 1, "TV 현재가치")
    wb.fcell(44, ord(f_term_col) - 64, f"={f_term_col}43*{last_col}15")
    wb.put(45, 1, "PV(FCFF 합계)")
    wb.fcell(45, 2, f"=SUM({f_fore_cols[0]}41:{last_col}41)")
    wb.put(46, 1, "PV(Terminal Value)")
    wb.fcell(46, 2, f"={f_term_col}44")
    wb.put(47, 1, "기업가치(EV)", bold=True)
    wb.fcell(47, 2, "=B45+B46", bold=True)
    wb.put(48, 1, "지분가치")
    wb.fcell(48, 2, "=B47-B8")
    wb.put(49, 1, "▶ 주당가치 (원)", bold=True)
    wb.fcell(49, 2, '=IF(B48<=0,"NM",B48/B6)', bold=True)

    wb.section(51, f"[차이 — FCFF({alt_mode}) 대비 FCFF({dcf_mode}) 기본값]")
    wb.put(52, 1, "EV 차이")
    wb.fcell(52, 2, "=B26-B47")
    wb.put(53, 1, "지분가치 차이")
    wb.fcell(53, 2, "=B28-B48")
    wb.put(54, 1, "주당가치 차이")
    wb.fcell(54, 2, '=IF(OR(ISTEXT(B29),ISTEXT(B49)),"n/a",B29-B49)')

    wb.section(56, "[Sensitivity Analysis — 주당가치(원), 계산 스냅샷]")
    wb.put(57, 1, "WACC\\g", bold=True)
    for j, g in enumerate(dcf_m["g_axis"]):
        wb.put(57, 2 + j, f"{g:.2f}%", bold=True)
    for i, w in enumerate(dcf_m["wacc_axis"]):
        r = 58 + i
        wb.put(r, 1, f"{w:.2f}%", bold=True)
        for j, g in enumerate(dcf_m["g_axis"]):
            v = dcf_m["sens_per_share"][w][g]
            wb.ws.cell(r, 2 + j, round(v) if v is not None else "n/a").number_format = "#,##0"

    sens2_hdr = 58 + len(dcf_m["wacc_axis"]) + 2
    wb.section(sens2_hdr - 1, "[Sensitivity Analysis — 기업가치 EV(원), 계산 스냅샷]")
    wb.put(sens2_hdr, 1, "WACC\\g", bold=True)
    for j, g in enumerate(dcf_m["g_axis"]):
        wb.put(sens2_hdr, 2 + j, f"{g:.2f}%", bold=True)
    for i, w in enumerate(dcf_m["wacc_axis"]):
        r = sens2_hdr + 1 + i
        wb.put(r, 1, f"{w:.2f}%", bold=True)
        for j, g in enumerate(dcf_m["g_axis"]):
            v = dcf_m["sens_ev"][w][g]
            wb.ws.cell(r, 2 + j, round(v) if v is not None else "n/a").number_format = "#,##0"

    wb.note(sens2_hdr + len(dcf_m["wacc_axis"]) + 3,
           "민감도표는 현재 가정 기준 계산 스냅샷(정적 값)입니다. 위 표들은 입력을 바꾸면 실시간 flex 됩니다.")

    # ── 시트5: Assumption Summary ──
    wb.new_sheet("Assumption Summary")
    wb.title(1, f"{m['company']} DCF Valuation 핵심 가정", None)
    wb.put(2, 1, "항목", bold=True); wb.put(2, 2, "가정값", bold=True)
    wb.put(2, 3, "단위", bold=True); wb.put(2, 4, "근거", bold=True)
    r = 4
    rows = [
        ("[수익 관련 가정]", None, None, None),
        ("매출원가율", f"{a['cogs_pct']:.2f}", "%", "최근 3개년 평균(실데이터 역산)"),
        ("판관비율", f"{a['sga_pct']:.2f}", "%", "최근 3개년 평균(실데이터 역산)"),
    ] + [
        (f"매출성장률({yr})", f"{g:.2f}", "%", "사용자 입력(compute_dcf 파라미터 — dcf_full_workbook 은 항상 명시값을 받음)")
        for yr, g in zip(m["fore_years"], a["growth"])
    ] + [
        ("[비용/투자 관련 가정]", None, None, None),
        ("법인세율", f"{a['tax_rate_pct']:.2f}", "%", "Damodaran 한국 법인세율(미지정 시)"),
        ("CAPEX율", f"{capex_pct:.2f}", "%(매출대비)", "사용자 입력"),
        ("운전자본 변동율", f"{nwc_pct:.2f}", "%(매출증가대비)", "사용자 입력"),
        ("신규 CAPEX 내용연수", f"{a['new_capex_life']}", "년", "유형자산/최근 D&A 역산(가능시), 아니면 10년 기본값"),
        ("[부채 관련 가정]", None, None, None),
        ("단기차입금 이자율", f"{a['short_rate_pct']:.2f}", "%", "실측 역산(가능시) 또는 시중금리 참고 기본값"),
        ("장기차입금 이자율", f"{a['long_rate_pct']:.2f}", "%", "시중금리 참고 기본값(조정 필요)"),
        ("[DCF Valuation 가정]", None, None, None),
        ("WACC", f"{wacc_pct:.2f}", "%", "사용자 입력"),
        ("영구성장률(g)", f"{terminal_growth_pct:.2f}", "%", "사용자 입력"),
        ("발행주식수", f"{dcf_m['shares']:,}", "주", "DART 실측"),
        ("FCFF 방식(기본)", dcf_mode, "raw|zero_floor",
         "기본값 raw(음수 FCFF를 0으로 감추지 않음) — DCF Valuation 시트에 대안 시나리오 병기"),
    ]
    for label, val, unit, rationale in rows:
        if val is None:
            wb.section(r, label)
        else:
            wb.put(r, 1, label); wb.put(r, 2, val); wb.put(r, 3, unit); wb.put(r, 4, rationale)
        r += 1

    r += 1
    wb.section(r, "[시트 간 수식 연동 구조]")
    r += 1
    for label, desc in [
        ("P&L and FCF", "히스토리(실측) + 추정치(가정 기반 수식)"),
        ("Debt Schedule", "이자비용 산출 → P&L and FCF 시트 이자비용(예측)에 연동"),
        ("Depreciation Schedule", "감가상각비 산출 → P&L and FCF 시트 D&A(예측)에 연동"),
        ("DCF Valuation", "P&L and FCF 의 FCFF(Raw, 기본) 참조 → TV·EV·주당가치 계산(Zero-floor는 대안 병기)"),
    ]:
        wb.put(r, 1, label); wb.put(r, 2, desc)
        r += 1

    r += 1
    wb.section(r, "[검증 결과 — build_full_model 스냅샷]")
    r += 1
    warnings_list = m["dcf"].get("warnings") or []
    if warnings_list:
        for w in warnings_list:
            wb.note(r, w)
            r += 1
    else:
        wb.note(r, "검증 통과 — 특이사항 없음(WACC>g, Terminal FCFF>0, 지분가치>0, TV 비중 정상범위).")
        r += 1

    r += 1
    wb.note(r, "⚠️ 개별 대출건 금리·자산별 상각내역·상환계획은 DART 구조화 데이터에 없어 가정입니다 — "
               "실제 조건을 알면 위 각 시트의 파란 셀을 직접 조정하세요.")

    fname = f"DCF_전체모델_{m['company']}_{m['as_of']}.xlsx".replace("/", "_")
    return wb.to_bytes(), fname


def _prev_col(c: str) -> str:
    return get_column_letter(ord(c[-1]) - 64 - 1) if len(c) == 1 else c


def _pl_col(m: dict, year: int) -> str:
    n_hist = len(m["hist_years"])
    idx = m["fore_years"].index(year)
    return get_column_letter(2 + n_hist + idx)


def _debt_col(m: dict, year: int) -> str:
    idx = m["fore_years"].index(year)
    return get_column_letter(4 + idx)


def _dep_col(m: dict, year: int) -> str:
    idx = m["fore_years"].index(year)
    return get_column_letter(4 + idx)
