"""FSI 원칙 기반 Excel 빌더.

- 입력(raw/가정) 셀 = 파란 글씨 + 셀 주석(출처/등급/기준일/공시일/URL)  ← 추적성·할루시네이션 방지
- 계산 셀 = 검정 글씨 + 살아있는 Excel 수식  ← 가정을 바꾸면 모델이 flex (formulas-over-hardcodes)
- '출처' 시트에 입력 셀의 소스를 자동 정리
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

from core.schema import Provenance, SourceType

_TIER_KO = {
    SourceType.AUTHORITATIVE: "공식(정부·거래소·중앙은행)",
    SourceType.PARSED_AUTHORITATIVE: "공시원문(직접 읽음 — 문서ID·인용 있음)",
    SourceType.REFERENCE: "참조(업계표준 데이터셋)",
    SourceType.COMPUTED: "계산(엔진)",
    SourceType.ASSUMPTION: "가정(사용자 입력)",
    SourceType.LLM_ESTIMATE: "LLM 추정 — 검증 필요",
}

_BLUE = Font(color="0000CC")            # 입력(하드코딩) — FSI 관례
_BLUE_B = Font(color="0000CC", bold=True)
_BLACK = Font(color="000000")
_BOLD = Font(bold=True)
_TITLE = Font(bold=True, size=14)
_HDR_FILL = PatternFill("solid", fgColor="1F3864")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_SECT_FILL = PatternFill("solid", fgColor="D9E1F2")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _prov_comment(prov: Provenance) -> Comment:
    lines = [
        f"출처: {prov.source}",
        f"등급: {_TIER_KO.get(prov.source_type, prov.source_type)}",
    ]
    if prov.as_of:
        lines.append(f"기준: {prov.as_of}")
    if prov.filing_date:
        lines.append(f"공시일: {prov.filing_date}")
    if prov.original_field:
        lines.append(f"필드: {prov.original_field}")
    if prov.source_url:
        lines.append(f"URL: {prov.source_url}")
    if prov.note:
        lines.append(f"비고: {prov.note}")
    c = Comment("\n".join(lines), "SKSQ Valuation Agent")
    c.width, c.height = 340, 170
    return c


class ValuationWorkbook:
    def __init__(self, sheet_title: str):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_title[:31]
        self.ws.column_dimensions["A"].width = 34
        self.ws.column_dimensions["B"].width = 26
        self.ws.column_dimensions["C"].width = 26
        self._sources: list[tuple[str, str, Provenance]] = []

    # ── 멀티시트 ─────────────────────────────────────────────
    def new_sheet(self, title: str, col_widths: dict | None = None):
        """새 시트를 만들고 활성 시트로 전환한다. 이후 label()/input()/formula() 등은
        전부 이 시트에 쓰인다 — 시트를 오가며 같은 메서드를 그대로 재사용."""
        self.ws = self.wb.create_sheet(title[:31])
        for col, w in (col_widths or {"A": 30, "B": 16}).items():
            self.ws.column_dimensions[col].width = w
        return self.ws

    def qref(self, sheet_title: str, coord: str) -> str:
        """다른 시트의 셀을 가리키는 수식 참조 문자열(예: "'Debt Schedule'!C23")."""
        return f"'{sheet_title[:31]}'!{coord}"

    # ── 셀 쓰기 ──────────────────────────────────────────────
    def title(self, row: int, text: str, subtitle: str | None = None):
        c = self.ws.cell(row, 1, text)
        c.font = _TITLE
        if subtitle:
            s = self.ws.cell(row + 1, 1, subtitle)
            s.font = Font(italic=True, color="C00000")

    def section(self, row: int, text: str):
        c = self.ws.cell(row, 1, text)
        c.font = _BOLD
        for col in (1, 2, 3):
            self.ws.cell(row, col).fill = _SECT_FILL

    def label(self, row: int, text: str, col: int = 1):
        self.ws.cell(row, col, text)

    def input(self, row: int, value, prov: Provenance, label: str | None = None,
              col: int = 2, fmt: str = "#,##0"):
        """파란 입력 셀 + 출처 주석. label 주면 A열에 라벨도 쓴다."""
        if label is not None:
            self.ws.cell(row, 1, label)
        c = self.ws.cell(row, col, value)
        c.font = _BLUE
        c.number_format = fmt
        c.border = _BORDER
        if prov is not None:
            c.comment = _prov_comment(prov)
            self._sources.append((c.coordinate, label or "", prov))
        return c.coordinate

    def formula(self, row: int, formula: str, label: str | None = None,
                col: int = 2, fmt: str = "#,##0", bold: bool = False):
        if label is not None:
            self.ws.cell(row, 1, label)
        c = self.ws.cell(row, col, formula)
        c.font = _BOLD if bold else _BLACK
        c.number_format = fmt
        c.border = _BORDER
        return c.coordinate

    def note(self, row: int, text: str):
        c = self.ws.cell(row, 1, text)
        c.font = Font(italic=True, size=9, color="808080")

    # ── 그리드(임의 위치) 헬퍼 ────────────────────────────────
    def put(self, row: int, col: int, text, bold: bool = False, italic: bool = False):
        c = self.ws.cell(row, col, text)
        if bold:
            c.font = _BOLD
        elif italic:
            c.font = Font(italic=True, color="808080")
        return c.coordinate

    def icell(self, row: int, col: int, value, prov: Provenance | None = None,
              fmt: str = "#,##0", src_label: str = "") -> str:
        c = self.ws.cell(row, col, value)
        c.font = _BLUE
        c.number_format = fmt
        c.border = _BORDER
        if prov is not None:
            c.comment = _prov_comment(prov)
            self._sources.append((c.coordinate, src_label, prov))
        return c.coordinate

    def fcell(self, row: int, col: int, formula: str, fmt: str = "#,##0",
              bold: bool = False) -> str:
        c = self.ws.cell(row, col, formula)
        c.font = _BOLD if bold else _BLACK
        c.number_format = fmt
        c.border = _BORDER
        return c.coordinate

    # ── 출처 시트 ────────────────────────────────────────────
    def build_sources_sheet(self):
        ws = self.wb.create_sheet("출처")
        headers = ["셀", "항목", "값 출처", "등급", "기준", "공시일", "URL"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h)
            c.fill, c.font = _HDR_FILL, _HDR_FONT
        for r, (coord, label, p) in enumerate(self._sources, 2):
            ws.cell(r, 1, coord)
            ws.cell(r, 2, label)
            ws.cell(r, 3, p.source)
            ws.cell(r, 4, _TIER_KO.get(p.source_type, p.source_type))
            ws.cell(r, 5, p.as_of or "")
            ws.cell(r, 6, p.filing_date or "")
            ws.cell(r, 7, p.source_url or "")
        for col, w in zip("ABCDEFG", (10, 24, 28, 26, 12, 12, 60)):
            ws.column_dimensions[col].width = w
        # 범례
        lr = len(self._sources) + 3
        ws.cell(lr, 1, "범례").font = _BOLD
        ws.cell(lr + 1, 1, "파란 글씨 = 입력(원천/가정), 검정 = 수식 계산")
        ws.cell(lr + 2, 1, "입력 셀에 마우스 올리면 출처 주석 표시. 입력을 바꾸면 수식이 자동 반영됩니다.")

    def to_bytes(self) -> bytes:
        self.build_sources_sheet()
        buf = BytesIO()
        self.wb.save(buf)
        return buf.getvalue()
